import os
import time
from pathlib import Path
import aiofiles
import asyncio
from aiogram import Bot
from aiogram.types import FSInputFile
from concurrent.futures import ThreadPoolExecutor

from openpyxl import load_workbook
from PyPDF2 import PdfMerger, PdfWriter, PdfReader
from fpdf import FPDF
from PIL import Image
from os.path import getsize
from io import BytesIO
from docx import Document

from app.data.config import MAX_SIZE_BYTES
from app.utils.clear_folder import async_clear_folder


async def merge_group(file_paths, output_path):
    merger = PdfMerger()
    for path in file_paths:
        merger.append(path)
    merger.write(output_path)
    merger.close()


async def group_and_merge_pdfs(pdf_paths: list[str], new_file_name: str,
                               output_dir: str = "temp_files") -> list[str]:
    current_group = []
    current_size = 0
    group_index = 1
    return_paths_list = []
    for i, path in enumerate(pdf_paths):
        size = await asyncio.to_thread(getsize, f"{path}")

        if size > MAX_SIZE_BYTES:
            if current_group:
                output_path = Path(output_dir) / f"{new_file_name}_Часть{group_index}.pdf"
                await merge_group(current_group, output_path)
                return_paths_list.append(output_path)
                group_index += 1

            output_path = Path(output_dir) / f"{new_file_name}_Часть{group_index}.pdf"
            await merge_group([path], output_path)
            return_paths_list.append(output_path)
            group_index += 1
            current_group = []
            current_size = 0

        elif current_size + size > MAX_SIZE_BYTES:
            output_path = Path(output_dir) / f"{new_file_name}_Часть{group_index}.pdf"
            await merge_group(current_group, output_path)
            return_paths_list.append(output_path)
            group_index += 1
            current_group = []
            current_size = 0

        current_group.append(path)
        current_size += size

    if current_group:
        output_path = Path(output_dir) / f"{new_file_name}_Часть{group_index}.pdf"
        await merge_group(current_group, output_path)
        return_paths_list.append(output_path)
    return return_paths_list


async def compress_file(input_file_path: str):
    output_file_path = input_file_path

    writer = PdfWriter()
    reader = PdfReader(input_file_path)

    for page in reader.pages:
        page.compress_content_streams()
        writer.add_page(page)

    buffer = BytesIO()
    writer.write(buffer)
    buffer.seek(0)

    async with aiofiles.open(output_file_path, 'wb') as f:
        await f.write(buffer.read())


async def txt_to_pdf(input_file_name, output_file_name) -> str:
    FONT_PATH = "DejaVuSans.ttf"

    async with aiofiles.open(input_file_name, 'r') as f:
        text = await f.read()
    pdf = FPDF()

    pdf.add_page()
    pdf.set_auto_page_break(auto=True, margin=15)

    # Добавляем пользовательский шрифт с поддержкой Unicode
    pdf.add_font("DejaVu", "", FONT_PATH, uni=True)
    pdf.set_font("DejaVu", size=12)
    pdf.multi_cell(0, 10, text=text)
    pdf.output(output_file_name)
    return output_file_name


async def image_to_pdf(input_file_name, output_file_name) -> str:
    image = Image.open(input_file_name)

    if image.mode in ("RGBA", "P"):
        image = image.convert("RGB")

    image.save(output_file_name, "PDF")
    return output_file_name


async def convert_xls_to_xlsx(input_path: str, output_dir: str) -> str:
    input_path = Path(input_path)

    output_dir = Path(output_dir)

    process = await asyncio.create_subprocess_exec(
        "libreoffice",
        "--headless",
        "--convert-to", "xlsx",
        "--outdir", str(output_dir),
        str(input_path),
        stdout=asyncio.subprocess.PIPE,
        stderr=asyncio.subprocess.PIPE,
    )

    stdout, stderr = await process.communicate()

    if process.returncode != 0:
        raise RuntimeError(f"Ошибка конвертации в .xlsx: {stderr.decode()}")

    output_xlsx = Path(output_dir) / (Path(input_path).stem + ".xlsx")
    return str(output_xlsx)


def _process_excel(file_path: str, output_path: str) -> str:
    """Синхронная часть: удаляет пустые строки и колонки из Excel-файла."""
    workbook = load_workbook(file_path)
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]

        # Удаление пустых строк
        for row in range(sheet.max_row, 0, -1):
            if all(cell.value is None for cell in sheet[row]):
                sheet.delete_rows(row)

        # Удаление пустых колонок с заголовками
        for col in range(sheet.max_column, 0, -1):
            column_data = [sheet.cell(row=row, column=col).value for row in range(2, sheet.max_row + 1)]
            if all(value is None for value in column_data):
                if sheet.cell(row=1, column=col).value:
                    print(f"Удаляю пустой столбец с заголовком: {sheet.cell(row=1, column=col).value}")
                sheet.delete_cols(col)

    cleaned_file = Path(output_path) / f"cleaned{int(time.time())}.xlsx"
    workbook.save(cleaned_file)
    return str(cleaned_file)


async def remove_empty_rows_and_cols(file_path: str, output_path: str) -> str:
    """Асинхронная обёртка для удаления пустых строк и колонок в Excel-файле."""
    loop = asyncio.get_event_loop()
    with ThreadPoolExecutor() as executor:
        result = await loop.run_in_executor(executor, _process_excel, file_path, output_path)
    return result


async def office_to_pdf(input_path: str, output_dir: str) -> str:
    input_path = Path(input_path)

    output_dir = Path(output_dir)

    process = await asyncio.create_subprocess_exec(
        "libreoffice",
        "--headless",
        "--convert-to", "pdf",
        "--outdir", str(output_dir),
        str(input_path),
        stdout=asyncio.subprocess.PIPE,
        stderr=asyncio.subprocess.PIPE,
    )

    stdout, stderr = await process.communicate()

    print("STDOUT:", stdout.decode())
    print("STDERR:", stderr.decode())
    output_pdf = output_dir / (Path(input_path).stem + ".pdf")
    return str(output_pdf)


async def get_convert(bot: Bot, files_ids_and_types: list[tuple[str, str, str]], new_file_name: str) -> list[FSInputFile] | tuple[str, str]:
    file_info = []

    file_names_for_clear = []
    pdf_files = []

    for i, (file_id, file_type, file_original_name) in enumerate(files_ids_and_types):
        file_names_for_clear.append(str(Path("temp_files") / f"{file_id}.{file_type}"))

        try:
            file = await bot.get_file(file_id)
            file_path = file.file_path
            output_dir = Path("temp_files")
            start_file_name = output_dir / f"{file_id}.{file_type}"
            output_filename = output_dir / f"{file_id}.pdf"
            output_filename, output_dir, start_file_name = str(output_filename), str(output_dir), str(start_file_name)
            await bot.download_file(file_path, start_file_name)
            if file_type == 'txt':
                output_filename = await txt_to_pdf(start_file_name, output_filename)
            elif file_type in ('jpg', 'png', 'jpeg'):
                output_filename = await image_to_pdf(start_file_name, output_filename)
            elif file_type in ('xls', 'xlsx'):
                xlsx_file = await convert_xls_to_xlsx(start_file_name, output_dir)
                file_names_for_clear.append(xlsx_file)
                cleaned_file = await remove_empty_rows_and_cols(xlsx_file, output_dir)
                file_names_for_clear.append(cleaned_file)
                output_filename = await office_to_pdf(cleaned_file, output_dir)
            elif file_type in ('doc', 'docx', 'ppt', 'pptx', 'rtf'):
                output_filename = await office_to_pdf(start_file_name, output_dir)

            if file_type != 'pdf':
                file_names_for_clear.append(output_filename)

            await compress_file(output_filename)

            pdf_files.append(output_filename)

            reader = PdfReader(output_filename)
            page_count = len(reader.pages)

            # Собираем информацию о файле
            file_info.append({
                'original_name': file_original_name,
                'page_count': page_count,
                'file_order': i + 1  # Порядковый номер файла
            })
        except Exception as e:
            print(e)
            # Удаляем временные файлы
            await async_clear_folder("", file_names_for_clear)

            return ('Ошибка', file_original_name)
    print(files_ids_and_types)
    paths = await group_and_merge_pdfs(pdf_paths=[file_path for file_path in pdf_files],
                                       new_file_name=new_file_name)

    doc = Document()
    
    # Создаем таблицу с заголовками
    table = doc.add_table(rows=1, cols=3)
    table.style = 'Table Grid'

    hdr_cells = table.rows[0].cells
    hdr_cells[0].text = 'Название'
    hdr_cells[1].text = 'Количество страниц'
    hdr_cells[2].text = 'Номер страницы'

    n = 1
    for info in file_info:
        filename = os.path.splitext(info['original_name'])[0]  # Убираем расширение
        row_cells = table.add_row().cells
        row_cells[0].text = filename
        row_cells[1].text = str(info['page_count'])
        row_cells[2].text = str(n)
        n += info['page_count']

    # Сохраняем DOCX в памяти
    docx_output = BytesIO()
    doc.save(docx_output)
    docx_output.seek(0)

    docx_filename = str(Path("temp_files") / f"{new_file_name}_info.docx")
    with open(docx_filename, 'wb') as f:
        f.write(docx_output.read())

    result_files: list[FSInputFile] = []
    if len(paths) == 1:
        result_files.append(FSInputFile(paths[0], filename=f'{new_file_name}.pdf'))
    else:
        for path in paths:
            result_files.append(FSInputFile(path))

    # Добавляем DOCX файл с информацией
    result_files.append(FSInputFile(docx_filename, filename=f'{new_file_name}_info.docx'))

    # Удаляем временные файлы, кроме конечных
    await async_clear_folder("", file_names_for_clear)

    return result_files
